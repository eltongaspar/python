# Pandas

import pandas as pd
import openpyxl 

pd.set_option("display.max_rows",None)
pd.set_option("display.min_rows",1)
pd.set_option("display.precision",5)

# Metodo 1 

Tab2021 = pd.read_excel("D:\Dados\Extrato_20210101_20211231.xlsx")
Tab2022 = pd.read_excel("D:\Dados\Extrato_20220101_20221231.xlsx")
Tab2023 = pd.read_excel("D:\Dados\Extrato_20230101_20231231.xlsx")
Tab2024 = pd.read_excel("D:\Dados\Extrato_20240101_20241231.xlsx")

print(Tab2021)
print(Tab2022)
print(Tab2023)
print(Tab2024)



#Metodo 2 
df2021 = pd.read_excel("D:\Dados\Extrato_20210101_20211231.xlsx")
df2022 = pd.read_excel("D:\Dados\Extrato_20220101_20221231.xlsx")
df2023 = pd.read_excel("D:\Dados\Extrato_20240101_20241231.xlsx")
df2024 = pd.read_excel("D:\Dados\Extrato_20240101_20241231.xlsx")


print(df2021)
print(df2022)
print(df2023)
print(df2024)

# Metodoe 3 
excel2021 = pd.read_excel("D:\Dados\Extrato_20210101_20211231.xlsx", sheet_name=None)
excel2022 = pd.read_excel("D:\Dados\Extrato_20220101_20221231.xlsx", sheet_name=None)
excel2023 = pd.read_excel("D:\Dados\Extrato_20240101_20241231.xlsx", sheet_name=None)
excel2024 = pd.read_excel("D:\Dados\Extrato_20240101_20241231.xlsx", sheet_name=None)

print(excel2021)
print(excel2022)
print(excel2023)
print(excel2024)
#print(df1['Planilha1.xlsx'])



wb2021 = openpyxl.load_workbook('D:\Dados\Extrato_20210101_20211231.xlsx')
wb2022 = openpyxl.load_workbook('D:\Dados\Extrato_20220101_20221231.xlsx')
wb2023 = openpyxl.load_workbook('D:\Dados\Extrato_20240101_20241231.xlsx')
wb2024 = openpyxl.load_workbook('D:\Dados\Extrato_20240101_20241231.xlsx')

ws2021 = (wb2021)
ws2022 = (wb2022)
ws2023 = (wb2023)
ws2024 = (wb2024)

print(ws2021)





