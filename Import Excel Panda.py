import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

tab2021 = pd.read_excel('D:\Dados\Extrato_20210101_20211231.xlsx')
tab2022 = pd.read_excel('D:\Dados\Extrato_20220101_20221231.xlsx')
tab2023 = pd.read_excel('D:\Dados\Extrato_20230101_20231231.xlsx')
tab2024 = pd.read_excel('D:\Dados\Extrato_20240101_20241231.xlsx')

excel_file = pd.read_excel('D:\Dados\Extrato_20210101_20211231.xlsx')
excel_file[['Categoria', 'Conta']]

report_table_2021 = excel_file.pivot_table(index='Categoria', 
                                      columns='Conta', 
                                      values='Valor', 
                                      aggfunc='sum').round(0)
print('2021')
print(report_table_2021)

excel_file = pd.read_excel('D:\Dados\Extrato_20220101_20221231.xlsx')
excel_file[['Categoria', 'Conta']]

report_table_2022 = excel_file.pivot_table(index='Categoria', 
                                      columns='Conta', 
                                      values='Valor', 
                                      aggfunc='sum').round(0)
print('2022')
print(report_table_2022)

excel_file = pd.read_excel('D:\Dados\Extrato_20230101_20231231.xlsx')
excel_file[['Categoria', 'Conta']]

report_table_2023 = excel_file.pivot_table(index='Categoria', 
                                      columns='Conta', 
                                      values='Valor', 
                                      aggfunc='sum').round(0)
print('2023')
print(report_table_2023)

excel_file = pd.read_excel('D:\Dados\Extrato_20240101_20241231.xlsx')
excel_file[['Categoria', 'Conta']]

report_table_2024 = excel_file.pivot_table(index='Categoria', 
                                      columns='Conta', 
                                      values='Valor', 
                                      aggfunc='sum').round(0)
print('2024')
print(report_table_2024)


report_table_2021.to_excel('report_2021.xlsx', 
                      sheet_name='Report', 
                      startrow=2)

report_table_2022.to_excel('report_2022.xlsx', 
                      sheet_name='Report', 
                      startrow=2)

report_table_2023.to_excel('report_2023.xlsx', 
                      sheet_name='Report', 
                      startrow=2)

report_table_2024.to_excel('report_2024.xlsx', 
                      sheet_name='Report', 
                      startrow=2)

wb2021 = load_workbook('report_2021.xlsx')
sheet2021 = wb2021['Report']

wb2022 = load_workbook('report_2022.xlsx')
sheet2022 = wb2022['Report']

wb2023 = load_workbook('report_2023.xlsx')
sheet2023 = wb2023['Report']

wb2024 = load_workbook('report_2024.xlsx')
sheet2024 = wb2024['Report']


# Referência para a planilha original
min_column = wb2021.active.min_column
max_column = wb2021.active.max_column
min_row = wb2021.active.min_row
max_row = wb2021.active.max_row

min_column = wb2022.active.min_column
max_column = wb2022.active.max_column
min_row = wb2022.active.min_row
max_row = wb2022.active.max_row

min_column = wb2023.active.min_column
max_column = wb2023.active.max_column
min_row = wb2023.active.min_row
max_row = wb2023.active.max_row

min_column = wb2024.active.min_column
max_column = wb2024.active.max_column
min_row = wb2024.active.min_row
max_row = wb2024.active.max_row