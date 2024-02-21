import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

def automate_excel(file_name):
    # para ler o arquivo em excel
    excel_file = pd.read_excel(file_name)
    # para fazer a Tabela Dinâmica
    report_table = excel_file.pivot_table(index='Categoria', columns='Conta', values='Valor', aggfunc='sum').round(0)
    # dividindo o mês e a extensão do nome do arquivo
    month_and_extension = file_name.split('_')[1]
    # enviar a tabela do relatório para um arquivo excel
    report_table.to_excel(f'report_{month_and_extension}', sheet_name='Report', startrow=2)
    
    # carregando pasta de trabalho e selecionando planilha
    wb = load_workbook(f'report_{month_and_extension}')
    sheet = wb['Report']
    # definindo as células de referência da planilha original
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    # realizando get no nome do mês
    month_name = month_and_extension.split('.')[0]
    #salvando
    wb.save(f'report_{month_and_extension}')
    return

automate_excel('D:\Dados\Extrato_20210101_20211231.xlsx')
#automate_excel('D:\Dados\Extrato_20220101_20221231.xlsxx')
#automate_excel('D:\Dados\Extrato_20230101_20231231.xlsx')
#automate_excel('D:\Dados\Extrato_20240101_20241231.xlsx')
